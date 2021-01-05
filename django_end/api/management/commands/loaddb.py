from django.core.management.base import BaseCommand, CommandError
from api.models import *
from openpyxl import load_workbook 


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print('กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A2'].value)
        print(f'count = {count}')
        #row4 = [ ws[f'{c}4'].value for c in 'ABCDEFG' ]
        #print( row4 )
        #column_names = ['id', 'rice_type']
        data = []  
        for i in range(count): # 0,1,2,3
            print(f'i = {i}')
            sheet_values = [ ws[f'{chr(65+j)}{4+i}'].value for j in range(len(column_names)) ]
            data.append( dict( (k,v) for k,v in zip(column_names, sheet_values)) )

        return data

    def handle(self, *args, **options):
        # pass
        from openpyxl import load_workbook
        filename = "xlsx/data.xlsx"
        wb = load_workbook(filename, data_only=True)
        #sheets = [ 'Rice_type', 'Work_status', 'Money_status', 'Work' ]

        print('กำลัง load ... Role')
        for d in self.load(wb, 'Role', ['id', 'role']):
            print(d)
            Role(**d).save()

        print('กำลัง load ... User')
        for d in self.load(wb, 'User', ['id', 'user_name','user_lastname', 'username','password' ,'email', 'phone' ,'avatar' ,'role' ,'lat' ,'lng']):
            User(**d).save()

        print('กำลัง load ... Mechanic_Type')
        for d in self.load(wb, 'Mechanic_Type', ['id', 'mechanic_type']):
            Mechanic_Type(**d).save()

        print('กำลัง load ... Mechanic')
        for d in self.load(wb, 'Mechanic', ['id', 'mechanic_detail','mechanic_img', 'mechanic_type']):
            Mechanic(**d).save()

        print('กำลัง load ... Store')
        for d in self.load(wb, 'Store', ['id', 'store_name','store_img', 'store_phone','adddres','lat','lng','time_open','time_close']):
            Store(**d).save()

        print('กำลัง load ... Admin')
        for d in self.load(wb, 'Admin', ['id', 'admin_name','admin_lastname', 'avatar','email','username','password']):
            Admin(**d).save()

        print('กำลัง load ... Product_Type')
        for d in self.load(wb, 'Product_Type', ['id', 'product_type']):
            Product_Type(**d).save()

        print('กำลัง load ... Product_Status')
        for d in self.load(wb, 'Product_Status', ['id', 'product_status']):
            Product_Status(**d).save()

        print('กำลัง load ... Product')
        for d in self.load(wb, 'Product', ['id', 'product_name','product_price', 'product_detail','product_img','product_type','product_status','amount']):
            Product(**d).save()

        print('กำลัง load ... Money_Status')
        for d in self.load(wb, 'Money_Status', ['id', 'money_status']):
            Money_Status(**d).save()
            
        print('กำลัง load ... Delivery_Options')
        for d in self.load(wb, 'Delivery_Options', ['id', 'delivery_options']):
            Delivery_Options(**d).save()

        print('กำลัง load ... Payment_Options')
        for d in self.load(wb, 'Payment_Options', ['id', 'payment_options']):
            Payment_Options(**d).save()

        print('กำลัง load ... Order')
        for d in self.load(wb, 'Order', ['id', 'date','user', 'admin','all_price','lat','lng','money_status','delivery_options','payment_options']):
            Order(**d).save()

        print('กำลัง load ... Payment')
        for d in self.load(wb, 'Payment', ['id', 'user','order', 'img','payment_options']):
            Payment(**d).save()

        print('กำลัง load ... Orderproduct')
        for d in self.load(wb, 'Orderproduct', ['id', 'order','orderproducts', 'amount']):
            Orderproduct(**d).save()

        print('กำลัง load ... Carts')
        for d in self.load(wb, 'Carts', ['id', 'user','product', 'amount']):
            Carts(**d).save()

        print('กำลัง load ... Conversations')
        for d in self.load(wb, 'Conversations', ['id', 'user','message', 'date','time']):
            Conversations(**d).save()

        print('กำลัง load ... Storck')
        for d in self.load(wb, 'Storck', ['id', 'all_products','Sold', 'inventories']):
            Storck(**d).save()


      